import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1F2D6E")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
EXAMPLE_FILL = PatternFill("solid", fgColor="E8F4FD")
EXAMPLE_FONT = Font(color="555555", italic=True)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN

def style_example(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN

def style_data(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN

# ─── Onglet 1 : Programme ─────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "1 - Programme"
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 20
ws1.row_dimensions[3].height = 20

headers1 = [
    "Titre du programme",
    "Sous-titre",
    "Première année",
    "Dernière année",
    "Date d'aujourd'hui (mois-année)"
]
for i, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=i, value=h)
    ws1.column_dimensions[get_column_letter(i)].width = 28

style_header(ws1, 1, len(headers1))

example1 = [
    "Programme de Transformation Finance",
    "Feuille de route prévisionnelle",
    2025,
    2027,
    "Juin 2025"
]
for i, v in enumerate(example1, 1):
    ws1.cell(row=2, column=i, value=v)
style_example(ws1, 2, len(headers1))

# ligne vide pour que l'utilisateur remplisse
for i in range(1, len(headers1) + 1):
    ws1.cell(row=3, column=i, value="")
style_data(ws1, 3, len(headers1))

# ─── Onglet 2 : Catégories ────────────────────────────────────────────────────
ws2 = wb.create_sheet("2 - Catégories")
ws2.row_dimensions[1].height = 30

headers2 = [
    "Nom de la catégorie",
    "Couleur (nom simple ou code hex)"
]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
    ws2.column_dimensions[get_column_letter(i)].width = 35

style_header(ws2, 1, len(headers2))

categories = [
    ("Performance Finance", "#1F2D6E"),
    ("Digitalisation & Automatisation", "#4472C4"),
    ("Data & Analytics", "#70AD47"),
    ("Systèmes & Technologie", "#A9D18E"),
    ("Organisation & Conduite du changement", "#ED7D31"),
]
for r, (nom, couleur) in enumerate(categories, 2):
    ws2.cell(row=r, column=1, value=nom)
    ws2.cell(row=r, column=2, value=couleur)
    ws2.row_dimensions[r].height = 20
    if r == 2:
        style_example(ws2, r, 2)
    else:
        style_data(ws2, r, 2)

# ─── Onglet 3 : Projets ───────────────────────────────────────────────────────
ws3 = wb.create_sheet("3 - Projets")
ws3.row_dimensions[1].height = 30

headers3 = [
    "Numéro du projet (#)",
    "Nom du projet",
    "Icône ou pictogramme",
    "Catégorie (doit correspondre à l'onglet Catégories)",
    "Chef de chantier",
    "Statut actuel (Terminé / En cours / À venir / En retard)"
]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
    ws3.column_dimensions[get_column_letter(i)].width = 30

style_header(ws3, 1, len(headers3))

projets = [
    (1, "Digitalisation de la facturation et automatisation", "Facture", "Digitalisation & Automatisation", "Marie Dupont", "En cours"),
    (2, "Optimisation du processus Order to Cash", "Panier", "Digitalisation & Automatisation", "Jean Martin", "En cours"),
    (3, "Renforcement du contrôle interne et conformité", "Bouclier", "Performance Finance", "Sophie Bernard", "En cours"),
]
for r, row in enumerate(projets, 2):
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
    ws3.row_dimensions[r].height = 22
    if r == 2:
        style_example(ws3, r, len(headers3))
    else:
        style_data(ws3, r, len(headers3))

# lignes vides
for r in range(len(projets) + 2, len(projets) + 17):
    for c in range(1, len(headers3) + 1):
        ws3.cell(row=r, column=c, value="")
    ws3.row_dimensions[r].height = 22
    style_data(ws3, r, len(headers3))

# ─── Onglet 4 : Phases (barres) ───────────────────────────────────────────────
ws4 = wb.create_sheet("4 - Phases des projets")
ws4.row_dimensions[1].height = 30

headers4 = [
    "Numéro du projet (#)",
    "Nom de la phase (texte dans la barre)",
    "Mois de début (ex: Janv. 2025)",
    "Mois de fin (ex: Mars 2025)"
]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
    ws4.column_dimensions[get_column_letter(i)].width = 35

style_header(ws4, 1, len(headers4))

phases = [
    (1, "Cadrage & conception", "Janv. 2025", "Mars 2025"),
    (1, "Build & paramétrage", "Avr. 2025", "Juil. 2025"),
    (1, "Déploiement multi-entités", "Nov. 2025", "Déc. 2027"),
    (2, "Analyse & design", "Janv. 2025", "Avr. 2025"),
    (2, "Pilotage & déploiement", "Mai 2025", "Oct. 2025"),
    (2, "Généralisation", "Janv. 2026", "Déc. 2027"),
]
for r, row in enumerate(phases, 2):
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
    ws4.row_dimensions[r].height = 22
    if r == 2:
        style_example(ws4, r, len(headers4))
    else:
        style_data(ws4, r, len(headers4))

for r in range(len(phases) + 2, len(phases) + 50):
    for c in range(1, len(headers4) + 1):
        ws4.cell(row=r, column=c, value="")
    ws4.row_dimensions[r].height = 22
    style_data(ws4, r, len(headers4))

# ─── Onglet 5 : Jalons des projets (losanges) ─────────────────────────────────
ws5 = wb.create_sheet("5 - Jalons des projets")
ws5.row_dimensions[1].height = 30

headers5 = [
    "Numéro du projet (#)",
    "Nom du jalon (ex: Mise en prod Phase 1)",
    "Date du jalon (ex: Oct. 2025)"
]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
    ws5.column_dimensions[get_column_letter(i)].width = 38

style_header(ws5, 1, len(headers5))

jalons_projets = [
    (1, "Mise en prod Phase 1", "Oct. 2025"),
    (3, "COPIL Contrôles", "Oct. 2025"),
    (4, "Recette utilisateurs", "Jan. 2026"),
    (8, "Recette Métier", "Jan. 2026"),
    (12, "Recette utilisateurs", "Jan. 2026"),
]
for r, row in enumerate(jalons_projets, 2):
    for c, val in enumerate(row, 1):
        ws5.cell(row=r, column=c, value=val)
    ws5.row_dimensions[r].height = 22
    if r == 2:
        style_example(ws5, r, len(headers5))
    else:
        style_data(ws5, r, len(headers5))

for r in range(len(jalons_projets) + 2, len(jalons_projets) + 30):
    for c in range(1, len(headers5) + 1):
        ws5.cell(row=r, column=c, value="")
    ws5.row_dimensions[r].height = 22
    style_data(ws5, r, len(headers5))

# ─── Onglet 6 : Jalons globaux (barre du bas) ─────────────────────────────────
ws6 = wb.create_sheet("6 - Jalons globaux")
ws6.row_dimensions[1].height = 30

headers6 = [
    "Nom du jalon global (ex: COPIL #1)",
    "Date (ex: Avr. 2025)"
]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=i, value=h)
    ws6.column_dimensions[get_column_letter(i)].width = 38

style_header(ws6, 1, len(headers6))

jalons_globaux = [
    ("Lancement programme", "Janv. 2025"),
    ("COPIL #1", "Avr. 2025"),
    ("COPIL #2", "Oct. 2025"),
    ("COPIL #3", "Fév. 2026"),
    ("COPIL #4", "Juin 2026"),
    ("COPIL #5", "Déc. 2026"),
    ("Bilan & perspectives", "T1 2027"),
]
for r, (nom, date) in enumerate(jalons_globaux, 2):
    ws6.cell(row=r, column=1, value=nom)
    ws6.cell(row=r, column=2, value=date)
    ws6.row_dimensions[r].height = 22
    if r == 2:
        style_example(ws6, r, 2)
    else:
        style_data(ws6, r, 2)

for r in range(len(jalons_globaux) + 2, len(jalons_globaux) + 15):
    ws6.cell(row=r, column=1, value="")
    ws6.cell(row=r, column=2, value="")
    ws6.row_dimensions[r].height = 22
    style_data(ws6, r, 2)

output = r"c:\cerveau 2 obsidian vault\ceveau 2 vault\data\roadmap_a_remplir.xlsx"
import os
os.makedirs(os.path.dirname(output), exist_ok=True)
wb.save(output)
print("Fichier créé :", output)
